# -*- coding: utf-8 -*-
"""디자인 보조 모드: 슬라이드 '스펙' → 브랜드 덱.

제작자가 내용·구성·강조를 정하고(편집 통제), 툴은 디자인만 한다.
스펙 예:
    @title: WordSense 성능 리포트
    @subtitle: 실험 결과 요약

    # 핵심 성능 지표
    > 어떤 조건에서도 높은 인식률
    - SNR -20dB에서 90%+ [강조]
    - 수백KB 초경량
    [차트: 막대, 데이터=88.5,90.1,91.4,92.3]

    # 방식 비교
    - A안 온디바이스 [강조]
    - B안 클라우드
    [차트: 원형, 데이터=A안:55, B안:30, 기타:15]
    [이미지: AI 음성인식 느낌, 배경, 투명20]

    # 데이터 연동 예
    [차트: 선, 데이터=../planing/sample_experiment_log.xlsx!실험기록!정확도(%)]
"""
import os
import re
import json
import subprocess
import tempfile

from src.common import paths

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = paths.RES_DIR
DEFAULT_ASSETS = paths.ASSETS_DIR
RENDERER = os.path.join(HERE, "render_deck.js")

_KIND = {"막대": "bar", "바": "bar", "bar": "bar", "원형": "pie", "파이": "pie", "pie": "pie",
         "선": "line", "라인": "line", "line": "line"}


def _resolve_excel(ref: str):
    """'path!sheet!column' 또는 'path!column' → (cats, vals)."""
    import openpyxl
    parts = [p.strip() for p in ref.split("!")]
    path = parts[0]
    if not os.path.isabs(path):
        # 스펙에 적힌 상대 경로는 사용자 작업 폴더 기준으로 푼다(리소스 폴더가 아니라).
        path = os.path.normpath(os.path.join(paths.DATA_DIR, path))
    wb = openpyxl.load_workbook(path, data_only=True)
    if len(parts) == 3:
        ws = wb[parts[1]] if parts[1] in wb.sheetnames else wb.worksheets[0]
        col_key = parts[2]
    else:
        ws = wb.worksheets[0]
        col_key = parts[1]
    headers = [str(c.value) for c in ws[1]]
    ci = next((i for i, h in enumerate(headers) if col_key.replace(" ", "") in h.replace(" ", "")), 2)
    cats, vals = [], []
    for r in range(2, ws.max_row + 1):
        c0 = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=ci + 1).value
        if v is None:
            continue
        cats.append(c0 if c0 is not None else r - 1)
        vals.append(v)
    return cats, vals


def _parse_chart(body: str):
    kind = "bar"
    for k, v in _KIND.items():
        if k in body:
            kind = v; break
    m = re.search(r"데이터\s*=\s*(.+)$", body)
    if not m:
        return None
    data = m.group(1).strip()
    if "!" in data and (".xlsx" in data or ".xls" in data):
        try:
            cats, vals = _resolve_excel(data)
            return {"kind": kind, "cats": cats, "vals": vals}
        except Exception as e:
            return {"kind": kind, "cats": ["오류"], "vals": [0], "error": str(e)}
    if ":" in data:  # label:value 쌍
        cats, vals = [], []
        for pair in data.split(","):
            if ":" in pair:
                lab, val = pair.split(":", 1)
                cats.append(lab.strip()); vals.append(float(re.sub(r"[^\d.\-]", "", val) or 0))
        return {"kind": kind, "cats": cats, "vals": vals}
    vals = [float(x) for x in re.findall(r"-?[\d.]+", data)]  # 값만
    return {"kind": kind, "cats": list(range(1, len(vals) + 1)), "vals": vals}


def _parse_image(body: str):
    parts = [p.strip() for p in body.split(",")]
    query = parts[0] if parts else ""
    mode = "background" if any("배경" in p for p in parts) else "region"
    tr = 0
    for p in parts:
        mt = re.search(r"투명\s*(\d+)", p)
        if mt:
            tr = int(mt.group(1))
    return {"query": query, "mode": mode, "transparency": tr, "path": None}


def build_plan(spec_text: str) -> dict:
    plan = {"title": "감바랩스 발표자료", "subtitle": "", "slides": [],
            "tagline": "Proprietary Technology, Ultimate Result",
            "contact": "Contact: contact@gambalabs.ai  |  https://gambalabs.ai"}
    cur = None
    for raw in spec_text.splitlines():
        line = raw.rstrip()
        s = line.strip()
        if not s:
            continue
        if s.lower().startswith("@title:"):
            plan["title"] = s.split(":", 1)[1].strip(); continue
        if s.lower().startswith("@subtitle:"):
            plan["subtitle"] = s.split(":", 1)[1].strip(); continue
        if s.startswith("# "):
            if cur:
                plan["slides"].append(cur)
            cur = {"type": "content", "title": s[2:].strip(), "lead": "", "bullets": []}
            continue
        if cur is None:
            continue
        if s.startswith("> "):
            cur["lead"] = s[2:].strip(); continue
        mc = re.match(r"\[차트\s*:\s*(.+)\]$", s)
        if mc:
            ch = _parse_chart(mc.group(1))
            if ch:
                cur["chart"] = ch
            continue
        mi = re.match(r"\[이미지\s*:\s*(.+)\]$", s)
        if mi:
            cur["image"] = _parse_image(mi.group(1)); continue
        if re.match(r"^[-*]\s+", s):
            txt = re.sub(r"^[-*]\s+", "", s)
            em = "[강조]" in txt
            txt = txt.replace("[강조]", "").strip()
            cur["bullets"].append({"text": txt, "emphasis": em})
            continue
        # 그 외 평문 → 리드 없으면 리드, 있으면 불릿
        if not cur["lead"]:
            cur["lead"] = s
        else:
            cur["bullets"].append({"text": s, "emphasis": False})
    if cur:
        plan["slides"].append(cur)
    return plan


def render_plan(plan: dict, out_pptx: str, assets_dir: str = DEFAULT_ASSETS, image_fetch=None,
                theme: str = None, theme_config: dict = None) -> str:
    """플랜(dict)을 그대로 렌더 — 스펙 파서/대화 에이전트가 공용으로 사용."""
    if theme:
        plan["theme"] = theme      # 내장 테마: 렌더러가 팔레트/에셋 선택
    if theme_config:
        plan["themeConfig"] = theme_config   # 커스텀 테마: 전체 설정 주입
    if image_fetch:  # 스톡 이미지 검색 훅
        for sl in plan.get("slides", []):
            if sl.get("image") and sl["image"].get("query"):
                sl["image"]["path"] = image_fetch(sl["image"]["query"])
    # 그라데이션 지시 → PNG 생성해 경로 주입(pptxgenjs 그라데이션 미지원 대체)
    for sl in plan.get("slides", []):
        g = sl.get("gradient")
        if isinstance(g, dict) and g.get("colors") and len(g["colors"]) >= 2:
            try:
                from app.ppt import gradient as _grad
                sl["gradientPath"] = _grad.gradient_png(
                    g["colors"][0], g["colors"][1], g.get("dir", "h"), bool(g.get("reverse")))
            except Exception as e:
                print(f"[render_plan] 그라데이션 생성 실패: {e}")
    os.makedirs(os.path.dirname(os.path.abspath(out_pptx)), exist_ok=True)
    with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as tf:
        json.dump(plan, tf, ensure_ascii=False)
        plan_path = tf.name
    try:
        res = subprocess.run([paths.node_exe(), RENDERER, plan_path, assets_dir, out_pptx],
                             capture_output=True, text=True, cwd=paths.node_cwd())
        if res.returncode != 0:
            raise RuntimeError(f"render_deck.js 실패: {res.stderr[-600:]}")
    finally:
        os.unlink(plan_path)
    return out_pptx


def generate(spec_text: str, out_pptx: str, assets_dir: str = DEFAULT_ASSETS, image_fetch=None,
             theme: str = None, theme_config: dict = None) -> str:
    return render_plan(build_plan(spec_text), out_pptx, assets_dir, image_fetch, theme, theme_config)


if __name__ == "__main__":
    spec = """@title: WordSense 성능 리포트
@subtitle: 실험 결과 요약

# 핵심 성능 지표
> 어떤 제약 조건에서도 높은 인식률과 구동성
- SNR -20dB에서 90%+ 인식률 [강조]
- 수백KB 초경량 모델, MCU 직접 탑재
- 100시간 오인식 0건
[차트: 막대, 데이터=88.5,90.1,91.4,92.3]

# 배포 방식 비교
- A안: 온디바이스 (권장) [강조]
- B안: 클라우드 의존
[차트: 원형, 데이터=온디바이스:55, 클라우드:30, 하이브리드:15]
[이미지: AI 음성인식 느낌, 배경, 투명15]

# 실험 정확도 추이 (Excel 연동)
> 기록된 실험 데이터로 자동 생성
[차트: 선, 데이터=../planing/sample_experiment_log.xlsx!실험기록!정확도(%)]
"""
    out = os.path.join(paths.OUTPUT_DIR, "spec_deck.pptx")
    print("생성:", generate(spec, out))
