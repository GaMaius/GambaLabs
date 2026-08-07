import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def create_template(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "영수증"

    # 기본 스타일
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # 칼럼 너비 조정
    for col in 'ABCDEFGHIJKLMN':
        ws.column_dimensions[col].width = 6

    # 제목
    ws.merge_cells('A1:N2')
    title_cell = ws['A1']
    title_cell.value = "영 수 증 (공급받는자 보관용)"
    title_cell.font = Font(size=20, bold=True)
    title_cell.alignment = center

    # 공급자 정보 (F4~M6 주변)
    ws.merge_cells('D4:E4')
    ws['D4'] = "사업자번호"
    ws['D4'].alignment = center
    ws['D4'].fill = header_fill
    ws.merge_cells('F4:N4')
    ws['F4'] = "" # 사업자번호 주입 위치

    ws.merge_cells('D5:E5')
    ws['D5'] = "상호"
    ws['D5'].alignment = center
    ws['D5'].fill = header_fill
    ws.merge_cells('F5:L5')
    ws['F5'] = "" # 상호 주입 위치

    ws['M5'] = "성명"
    ws['M5'].fill = header_fill
    ws['M5'].alignment = center
    ws['N5'] = "" # 성명 주입 위치

    ws.merge_cells('D6:E6')
    ws['D6'] = "소재지"
    ws['D6'].alignment = center
    ws['D6'].fill = header_fill
    ws.merge_cells('F6:N6')
    ws['F6'] = "" # 주소 주입 위치

    # 작성일자
    ws.merge_cells('B8:D8')
    ws['B8'] = "" # 일자 주입 위치
    ws['B8'].alignment = center

    # 품목 리스트 헤더 (11행)
    headers = [("A11:B11", "월일"), ("C11:F11", "품목"), ("G11:H11", "수량"), ("I11:K11", "단가"), ("L11:N11", "금액")]
    for cell_range, text in headers:
        ws.merge_cells(cell_range)
        start_cell = cell_range.split(':')[0]
        ws[start_cell] = text
        ws[start_cell].alignment = center
        ws[start_cell].fill = header_fill

    # 품목 입력란 (12행 ~ 26행)
    for row in range(12, 27):
        ws.merge_cells(f'A{row}:B{row}')
        ws.merge_cells(f'C{row}:F{row}') # 품목명: C열
        ws.merge_cells(f'G{row}:H{row}') # 수량: G열
        ws.merge_cells(f'I{row}:K{row}') # 단가: I열
        ws.merge_cells(f'L{row}:N{row}') # 금액: L열

        # 금액 수식 (수량 * 단가)
        ws[f'L{row}'] = f"=G{row}*I{row}"
        
        # 숫자 형식 포맷
        ws[f'I{row}'].number_format = '#,##0'
        ws[f'L{row}'].number_format = '#,##0'

    # 합계
    ws.merge_cells('A27:K27')
    ws['A27'] = "합 계"
    ws['A27'].alignment = center
    ws['A27'].fill = header_fill
    ws.merge_cells('L27:N27')
    ws['L27'] = "=SUM(L12:L26)"
    ws['L27'].number_format = '#,##0'
    ws['L27'].font = Font(bold=True)

    # 테두리 적용
    for row in ws.iter_rows(min_row=4, max_row=6, min_col=4, max_col=14):
        for cell in row:
            cell.border = thin_border
            
    for row in ws.iter_rows(min_row=11, max_row=27, min_col=1, max_col=14):
        for cell in row:
            cell.border = thin_border

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    print(f"Template created at {path}")

if __name__ == "__main__":
    create_template(r"c:\Users\rkfka\Desktop\test\gambalabs-office-automation\assets\receipt_template.xlsx")
