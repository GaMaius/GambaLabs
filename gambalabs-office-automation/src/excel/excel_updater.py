import os
import re
import openpyxl
from typing import List, Dict, Any

from src.excel.table_utils import detect_table

class ExcelUpdater:
    """
    openpyxl 기반 수식, 조건부 서식, 차트를 100% 보존하며 셀 단위 데이터 업데이트를 처리하는 엔진
    """
    @staticmethod
    def _is_formula(value: Any) -> bool:
        """대상 셀이 수식이면 True — 실수로 수식을 덮어쓰지 않도록 보호."""
        return isinstance(value, str) and value.startswith("=")

    def apply_updates(self, excel_path: str, updates: List[Dict[str, Any]], output_path: str) -> List[str]:
        """
        excel_path 파일을 읽어 updates 내역을 반영하고 output_path에 저장.
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")

        # data_only=False로 수식 원본 100% 보존
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        change_logs = []

        for update in updates:
            sheet_name = update.get("sheet_name")
            item_name = update.get("item_name")
            column_name = update.get("column_name")
            cell_address = update.get("cell_address")
            new_value = update.get("new_value")

            ws = None
            # 시트명 매칭
            for sname in wb.sheetnames:
                if sheet_name in sname or sname in sheet_name:
                    ws = wb[sname]
                    break

            if ws is None:
                change_logs.append(f"[SKIP] 시트를 찾을 수 없음: {sheet_name}")
                continue

            # 1. 셀 주소가 지정된 경우 직접 업데이트
            if cell_address:
                old_val = ws[cell_address].value
                if self._is_formula(old_val):
                    change_logs.append(f"[SKIP] 수식 셀 보호: {ws.title}!{cell_address} ('{old_val}')")
                    continue
                ws[cell_address] = new_value
                change_logs.append(f"[{ws.title}] 셀 {cell_address}: '{old_val}' -> '{new_value}'")
                continue

            # 2. 항목명과 컬럼명으로 동적 셀 탐색
            target_row = None
            target_col = None

            # 헤더 행 탐색 — 1행이 아닐 수 있어 자동 감지(제목·빈 줄 뒤 표 지원)
            hr, header_row, first_data = detect_table(ws)
            clean_target_col = re.sub(r"[\s()원]", "", str(column_name))
            for col_idx, h_val in enumerate(header_row, start=1):
                if h_val:
                    clean_h_val = re.sub(r"[\s()원]", "", str(h_val))
                    if clean_target_col in clean_h_val or clean_h_val in clean_target_col:
                        target_col = col_idx
                        break

            # 항목 행(헤더 아래 데이터 영역) 탐색
            for row_idx in range(first_data, ws.max_row + 1):
                cell_item = ws.cell(row=row_idx, column=1).value
                if cell_item and (item_name in str(cell_item) or str(cell_item) in item_name):
                    target_row = row_idx
                    break

            if target_row and target_col:
                target_cell = ws.cell(row=target_row, column=target_col)
                old_val = target_cell.value
                if self._is_formula(old_val):
                    change_logs.append(f"[SKIP] 수식 셀 보호: {ws.title}!{target_cell.coordinate} ('{old_val}')")
                    continue
                target_cell.value = new_value
                cell_coord = target_cell.coordinate
                change_logs.append(f"[{ws.title}] 셀 {cell_coord} ({item_name} -> {column_name}): '{old_val}' -> '{new_value}'")
            else:
                change_logs.append(f"[SKIP] 항목 또는 컬럼 위치 탐색 실패 (항목: {item_name}, 컬럼: {column_name})")

        # 저장
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        print(f"[ExcelUpdater] 성공적으로 엑셀 수정 및 저장 완료: {output_path}")
        return change_logs
