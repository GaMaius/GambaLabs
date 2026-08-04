# -*- coding: utf-8 -*-
"""자연어 수정 요청 → 엑셀 셀 변경 목록(JSON) 추출기.

- LLM 백엔드는 OpenAI 호환 인터페이스로 스왑 가능:
    OPENAI_API_KEY   : 키
    OPENAI_BASE_URL  : 엔드포인트 (미설정=OpenAI, 사내/로컬은 http://host:11434/v1 등)
    OPENAI_MODEL     : 모델명 (기본 gpt-4o)
- 실제 워크북 구조(시트·헤더·항목)를 프롬프트에 넣어 특정 샘플에 하드코딩되지 않게 일반화.
- 키가 없으면 워크북 구조 기반 휴리스틱으로 best-effort 폴백(범용 LLM 대비 제한적).
"""
import os
import re
import json
from typing import List, Optional, Dict, Any

import openpyxl
from pydantic import BaseModel, Field


class CellUpdate(BaseModel):
    sheet_name: str = Field(description="변경 대상 시트명 (워크북에 실제 존재하는 이름)")
    item_name: str = Field(description="변경 대상 항목(행) 이름 (1열 값과 매칭)")
    column_name: str = Field(description="변경 대상 컬럼(열) 이름 (헤더행과 매칭)")
    new_value: Any = Field(description="최종 반영 값 (합산/연장 등 계산이 필요하면 계산 후 최종값)")
    reason: Optional[str] = Field(default=None, description="근거(선택)")


def build_workbook_context(excel_path: str, max_items: int = 60) -> str:
    """시트별 헤더·1열 항목 목록을 요약해 LLM 프롬프트용 문맥 문자열 생성.
    헤더가 1행이 아닌 시트도 자동 감지한다."""
    from src.excel.table_utils import detect_table
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    lines: List[str] = []
    for ws in wb.worksheets:
        hr, header_vals, first_data = detect_table(ws)
        headers = [str(v) for v in header_vals if v is not None]
        items: List[str] = []
        for r in range(first_data, min(ws.max_row, first_data + max_items - 1) + 1):
            v = ws.cell(row=r, column=1).value
            if v is not None and str(v).strip():
                items.append(str(v))
        lines.append(f"- 시트 '{ws.title}': 헤더={headers}; 1열 항목={items}")
    return "\n".join(lines)


class LLMExtractor:
    def __init__(self, api_key: Optional[str] = None, base_url: Optional[str] = None, model: Optional[str] = None):
        self.api_key = api_key or os.getenv("OPENAI_API_KEY")
        self.base_url = base_url or os.getenv("OPENAI_BASE_URL")  # None → OpenAI 기본
        self.model = model or os.getenv("OPENAI_MODEL", "gpt-4o")

    def extract_updates(self, request_text: str, excel_path: Optional[str] = None) -> List[Dict[str, Any]]:
        context = build_workbook_context(excel_path) if excel_path else ""
        if self.api_key or self.base_url:
            try:
                return self._extract_with_llm(request_text, context)
            except Exception as e:
                print(f"[LLMExtractor] LLM 호출 실패 ({e}). 휴리스틱 폴백으로 전환합니다.")
        else:
            print("[LLMExtractor] OPENAI_API_KEY/BASE_URL 미설정 → 휴리스틱 폴백(제한적).")
        return self._heuristic_fallback(request_text, excel_path)

    # ── LLM 경로 (OpenAI 호환) ─────────────────────────────
    def _extract_with_llm(self, text: str, context: str) -> List[Dict[str, Any]]:
        from openai import OpenAI

        client = OpenAI(api_key=self.api_key or "not-needed", base_url=self.base_url)
        prompt = f"""너는 엑셀 대시보드 업데이트 도우미다. 아래 워크북 구조와 사용자 요청을 보고,
변경해야 할 셀들을 JSON으로만 반환하라.

[워크북 구조]
{context}

[사용자 요청]
\"\"\"{text}\"\"\"

[반환 형식] 오직 아래 JSON만:
{{"updates": [{{"sheet_name": "...", "item_name": "...", "column_name": "...", "new_value": ...}}]}}

[규칙]
- sheet_name/column_name/item_name 은 반드시 위 워크북 구조에 실제로 존재하는 이름으로 쓸 것.
- 금액 합산/일정 연장 등 계산이 필요하면 계산해서 new_value 에 최종값만 넣을 것(예: 기존 150만 + 200만 = 3500000).
- 날짜는 YYYY-MM-DD, 금액은 정수(콤마 없이)."""
        resp = client.chat.completions.create(
            model=self.model,
            temperature=0,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": "You output only valid JSON. No prose."},
                {"role": "user", "content": prompt},
            ],
        )
        data = json.loads(resp.choices[0].message.content)
        return [CellUpdate(**u).model_dump() for u in data.get("updates", [])]

    # ── 오프라인 휴리스틱 폴백 (자연어 채팅 대응, best-effort) ──
    @staticmethod
    def _parse_dates(text: str) -> List[str]:
        out = re.findall(r"\d{4}-\d{2}-\d{2}", text)
        for mo, d in re.findall(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일", text):
            out.append(f"2026-{int(mo):02d}-{int(d):02d}")
        for mo, d in re.findall(r"(?<!\d)(\d{1,2})/(\d{1,2})(?!\d)", text):
            out.append(f"2026-{int(mo):02d}-{int(d):02d}")
        return out

    @staticmethod
    def _parse_amounts(text: str) -> List[int]:
        out = []
        for n in re.findall(r"([\d,]+)\s*만\s*원?", text):   # "350만원", "350만"
            out.append(int(n.replace(",", "")) * 10000)
        for n in re.findall(r"([\d,]{5,})\s*원", text):        # "3,500,000원"
            out.append(int(n.replace(",", "")))
        return out

    @staticmethod
    def _tokens(s: str) -> List[str]:
        return [t for t in re.split(r"[\s/·,()'\"]+", s) if len(t) >= 2 and t not in ("및", "the", "of")]

    def _heuristic_fallback(self, text: str, excel_path: Optional[str]) -> List[Dict[str, Any]]:
        if not excel_path:
            return []
        from src.excel.table_utils import detect_table
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        updates: List[Dict[str, Any]] = []

        dates = self._parse_dates(text)
        amounts = self._parse_amounts(text)
        status_words = [w for w in ["지연", "완료", "진행중", "대기", "보류", "중단"] if w in text]

        for ws in wb.worksheets:
            _hr, header_vals, first_data = detect_table(ws)
            headers = [str(v) for v in header_vals if v is not None]
            # 항목명 토큰이 요청문과 가장 많이 겹치는 행 1개 선택(자연어 축약 대응)
            best_score, best_item = 0, None
            for r in range(first_data, ws.max_row + 1):
                item = ws.cell(row=r, column=1).value
                if not item:
                    continue
                score = sum(1 for t in self._tokens(str(item)) if t in text)
                if score > best_score:
                    best_score, best_item = score, str(item)
            if best_score < 2 or best_item is None:   # 단일 토큰 오탐 방지
                continue
            for hname in headers:
                key = re.sub(r"[\s()원]", "", hname)
                if "종료" in key and dates:
                    updates.append({"sheet_name": ws.title, "item_name": best_item, "column_name": hname, "new_value": dates[-1]})
                elif "상태" in key and status_words:
                    updates.append({"sheet_name": ws.title, "item_name": best_item, "column_name": hname, "new_value": status_words[-1]})
                elif "금액" in key and amounts:
                    updates.append({"sheet_name": ws.title, "item_name": best_item, "column_name": hname, "new_value": max(amounts)})
        return updates
