# -*- coding: utf-8 -*-
"""실험/재무 결과 기록 + 자동 차트 엔진.

자연어 결과 설명("4회차 정확도 92.3%, SNR -20dB, 오인식 1건")을 트래커 시트에
한 줄로 기록하고, 추이 차트를 데이터 증가에 맞춰 자동으로 다시 그린다.
수식이 있는 셀은 건드리지 않는다(수식 보호는 excel_updater와 동일 철학).
"""
import os
import re
import json
import datetime
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.chart import LineChart, Reference


class ExperimentLogger:
    def __init__(self, sheet_name: Optional[str] = None):
        self.sheet_name = sheet_name

    def _sheet(self, wb):
        if self.sheet_name and self.sheet_name in wb.sheetnames:
            return wb[self.sheet_name]
        return wb.worksheets[0]

    def parse_result(self, text: str, headers: List[str], next_round: int) -> List[Any]:
        row = []
        for h in headers:
            hs = str(h or "")
            key = re.sub(r"\(.*?\)|\s|%|건|회", "", hs)  # 단위 제거
            if "회차" in hs:
                m = re.search(r"(\d+)\s*회", text)
                row.append(int(m.group(1)) if m else next_round)
            elif "날짜" in hs or "일자" in hs:
                row.append(datetime.date.today().isoformat())
            elif key and re.search(re.escape(key) + r"\s*(-?[\d.]+)", text):
                v = re.search(re.escape(key) + r"\s*(-?[\d.]+)", text).group(1)
                row.append(float(v) if "." in v else int(v))
            else:
                # 헤더명이 텍스트에 없으면: 비고류엔 원문 일부, 그 외 공란
                row.append("" if ("비고" in hs or "메모" in hs) else None)
        return row

    @staticmethod
    def _llm_available() -> bool:
        return bool(os.getenv("OPENAI_API_KEY") or os.getenv("OPENAI_BASE_URL"))

    def parse_result_llm(self, text: str, headers: List[Any], next_round: int) -> List[Any]:
        """자연어 결과 설명을 LLM이 트래커 헤더에 맞춰 한 행으로 변환. 실패 시 정규식 파싱으로 폴백."""
        from openai import OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY") or "ollama", base_url=os.getenv("OPENAI_BASE_URL"))
        hlist = [str(h) for h in headers]
        prompt = f"""실험/재무 결과 설명을 트래커 표의 '한 행'으로 변환하라. 아래 컬럼 순서·이름에 정확히 맞춰라.

[컬럼] {hlist}
[설명] \"\"\"{text}\"\"\"

[규칙]
- 각 컬럼 값을 설명에서 찾아 넣어라. 숫자는 단위·기호 빼고 숫자만 (예: '정확도(%)' → 92.3, 'SNR(dB)' → -20, '오인식(건)' → 1).
- '회차'가 설명에 없으면 {next_round}. '날짜/일자'가 없으면 오늘({datetime.date.today().isoformat()}).
- 설명에 근거가 없는 수치 컬럼은 빈 문자열 "". '비고/메모'류엔 설명의 핵심을 10자 내외로.
- 오직 JSON만: {{"row": {{"<컬럼명>": <값>, ...}}}}"""
        resp = client.chat.completions.create(
            model=os.getenv("OPENAI_MODEL", "gpt-4o"), temperature=0,
            response_format={"type": "json_object"},
            messages=[{"role": "system", "content": "You output only valid JSON. 한국어 값은 그대로 보존."},
                      {"role": "user", "content": prompt}])
        rowmap = (json.loads(resp.choices[0].message.content) or {}).get("row", {})
        rowmap_clean = {self._clean(k): v for k, v in rowmap.items()}
        out = []
        for h in headers:
            hs = str(h or "")
            val = rowmap.get(hs, None)
            if val in (None, ""):
                val = rowmap_clean.get(self._clean(hs), None)
            if val in (None, ""):
                if "회차" in hs:
                    val = next_round
                elif "날짜" in hs or "일자" in hs:
                    val = datetime.date.today().isoformat()
                else:
                    val = ""
            out.append(val)
        return out

    def parse(self, text: str, headers: List[Any], next_round: int, use_llm: bool = False) -> List[Any]:
        """use_llm=True고 LLM이 설정돼 있으면 LLM 파싱, 아니면(또는 실패 시) 정규식 파싱."""
        if use_llm and self._llm_available():
            try:
                return self.parse_result_llm(text, headers, next_round)
            except Exception as e:
                print(f"[ExperimentLogger] LLM 파싱 실패({e}) → 규칙 기반")
        return self.parse_result(text, headers, next_round)

    _PALETTE = ["0038A3", "2A9D99", "E5484D", "D9A400", "6E56CF", "1970AE"]

    def _rebuild_chart(self, ws, n_data: int, value_col: int = 3, cat_col: int = 1,
                       anchor: str = "H2", value_cols: List[int] = None):
        """추이 라인차트 재생성. value_cols가 주어지면 다중 지표(시리즈)로 그린다."""
        ws._charts = []  # 기존 차트 제거 후 재생성(범위 자동 확장)
        chart = LineChart()
        cols = [c for c in (value_cols or [value_col]) if c]
        chart.title = "지표 추이" if len(cols) > 1 else "추이"
        chart.height, chart.width = 8, 18
        chart.y_axis.majorGridlines = None
        cats = Reference(ws, min_col=cat_col, min_row=2, max_row=1 + n_data)
        for c in cols:
            data = Reference(ws, min_col=c, min_row=1, max_row=1 + n_data)
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        for i, ser in enumerate(chart.series):
            ser.graphicalProperties.line.solidFill = self._PALETTE[i % len(self._PALETTE)]
            ser.graphicalProperties.line.width = 28000
            ser.smooth = False
        if len(cols) > 1:
            chart.legend.position = "b"
        else:
            chart.legend = None
        ws.add_chart(chart, anchor)

    def _numeric_value_cols(self, ws, headers) -> List[int]:
        """차트에 쓸 수치 지표 열 목록(회차·날짜·비고 등 키/텍스트 열 제외)."""
        SKIP = ("회차", "날짜", "일자", "비고", "메모", "구분", "이름", "항목")
        r = ws.max_row if ws.max_row >= 2 else 2
        cols = []
        for ci in range(2, len(headers) + 1):
            hs = str(headers[ci - 1] or "")
            if any(k in hs for k in SKIP):
                continue
            v = ws.cell(row=r, column=ci).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cols.append(ci)
        return cols

    def create_tracker(self, path: str, headers: List[str], title: str = "기록") -> Dict[str, Any]:
        """지정한 헤더로 빈 트래커 워크북 생성(첫 기록부터 차트가 붙도록 준비)."""
        headers = [str(h).strip() for h in headers if str(h).strip()]
        if not headers:
            raise ValueError("헤더가 비었습니다.")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (title or "기록")[:31]
        ws.append(headers)
        from openpyxl.styles import Font, PatternFill
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0038A3")
        for i, _h in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        wb.save(path)
        return {"out": path, "headers": headers, "sheet": ws.title}

    @staticmethod
    def _clean(s: str) -> str:
        return re.sub(r"\(.*?\)|\s|%|건|회", "", str(s or ""))

    def _row_from_dict(self, headers, data: Dict[str, Any], next_round: int):
        clean_data = {self._clean(k): v for k, v in data.items()}
        row = []
        for h in headers:
            hs = str(h or "")
            if "회차" in hs:
                row.append(data.get(h) or clean_data.get(self._clean(h)) or next_round)
            elif "날짜" in hs or "일자" in hs:
                row.append(data.get(h) or clean_data.get(self._clean(h)) or datetime.date.today().isoformat())
            else:
                row.append(clean_data.get(self._clean(h), ""))
        return row

    @staticmethod
    def _value_col_index(ws, headers) -> int:
        """차트에 쓸 수치 컬럼(첫 숫자형 열, 회차/날짜/비고 등 키·텍스트 열 제외)."""
        SKIP = ("회차", "날짜", "일자", "비고", "메모", "구분", "이름", "항목")
        # 마지막 데이터 행 기준으로 판정(방금 추가한 행에 값이 있을 수 있음)
        r = ws.max_row if ws.max_row >= 2 else 2
        for ci in range(2, len(headers) + 1):
            hs = str(headers[ci - 1] or "")
            if any(k in hs for k in SKIP):
                continue
            v = ws.cell(row=r, column=ci).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return ci
        # 폴백: 이름으로 못 거르면 첫 숫자형 열
        for ci in range(2, len(headers) + 1):
            v = ws.cell(row=r, column=ci).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return ci
        return min(3, len(headers))

    def append_row(self, xlsx_path: str, data: Dict[str, Any], output_path: str = None) -> Dict[str, Any]:
        """구조화된 dict를 트래커에 한 행 추가(파서 없이). 스크립트 연동용."""
        wb = openpyxl.load_workbook(xlsx_path)
        ws = self._sheet(wb)
        headers = [c.value for c in ws[1]]
        row = self._row_from_dict(headers, data, ws.max_row)
        ws.append(row)
        n_data = ws.max_row - 1
        self._rebuild_chart(ws, n_data, value_col=self._value_col_index(ws, headers))
        out = output_path or xlsx_path  # 기본은 제자리 업데이트(누적)
        wb.save(out)
        return {"sheet": ws.title, "appended": dict(zip([str(h) for h in headers], row)), "total_rows": n_data, "out": out}

    def preview_result(self, xlsx_path: str, description: str, use_llm: bool = False) -> Dict[str, Any]:
        """저장하지 않고, 자연어 설명이 어떤 한 행으로 기록될지 미리 보여준다."""
        wb = openpyxl.load_workbook(xlsx_path)
        ws = self._sheet(wb)
        headers = [c.value for c in ws[1]]
        data_rows = ws.max_row - 1
        row = self.parse(description, headers, next_round=data_rows + 1, use_llm=use_llm)
        hstr = [str(h) for h in headers]
        blanks = [hstr[i] for i, v in enumerate(row) if v in (None, "")]
        return {"sheet": ws.title, "headers": hstr, "row": row,
                "mapped": dict(zip(hstr, row)), "blanks": blanks,
                "engine": ("AI 파싱" if (use_llm and self._llm_available()) else "규칙 기반")}

    def append_result(self, xlsx_path: str, description: str, output_path: str,
                      value_col: int = None, use_llm: bool = False,
                      chart_mode: str = "auto") -> Dict[str, Any]:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = self._sheet(wb)
        headers = [c.value for c in ws[1]]
        data_rows = ws.max_row - 1
        row = self.parse(description, headers, next_round=data_rows + 1, use_llm=use_llm)
        ws.append(row)
        n_data = ws.max_row - 1
        if chart_mode == "all":
            vcols = self._numeric_value_cols(ws, headers) or [self._value_col_index(ws, headers)]
            self._rebuild_chart(ws, n_data, value_cols=vcols)
            metric = ", ".join(str(headers[c - 1]) for c in vcols if 0 < c <= len(headers))
        else:
            vc = value_col or self._value_col_index(ws, headers)   # 하드코딩 대신 자동 감지
            self._rebuild_chart(ws, n_data, value_col=vc)
            metric = str(headers[vc - 1]) if 0 < vc <= len(headers) else ""
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        return {"sheet": ws.title, "appended": dict(zip([str(h) for h in headers], row)),
                "total_rows": n_data, "out": output_path, "chart_metric": metric,
                "engine": ("AI 파싱" if (use_llm and self._llm_available()) else "규칙 기반")}


def create_sample(path: str):
    """데모용 실험 트래커 생성(데이터 + 라인차트)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "실험기록"
    ws.append(["회차", "날짜", "정확도(%)", "SNR(dB)", "오인식(건)", "비고"])
    rows = [
        [1, "2026-07-01", 88.5, -10, 0, "초기 모델"],
        [2, "2026-07-08", 90.1, -15, 1, "하이퍼파라미터 튜닝"],
        [3, "2026-07-15", 91.4, -18, 0, "데이터 증강"],
    ]
    for r in rows:
        ws.append(r)
    ExperimentLogger()._rebuild_chart(ws, len(rows))
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    return path


if __name__ == "__main__":
    import sys
    base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    sample = os.path.join(base, "planing", "sample_experiment_log.xlsx")
    if not os.path.exists(sample) or "--recreate" in sys.argv:
        print("샘플 생성:", create_sample(sample))
    out = os.path.join(base, "output", "experiment_logged.xlsx")
    res = ExperimentLogger().append_result(sample, "4회차 정확도 92.3%, SNR -20dB, 오인식 1건, 노이즈 강인성 개선", out)
    import json
    print(json.dumps(res, ensure_ascii=False, indent=2))
