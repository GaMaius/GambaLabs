# -*- coding: utf-8 -*-
"""gamba_log 실사용 시나리오 테스트.

실제 실험/연구 과정에서 발생하는 지저분한 상황들을 재현한다:
- 트래커 파일이 없는 상태에서 첫 기록 (cold start)
- 키 일부만 있는 불완전한 기록
- 트래커에 없는 새 키가 섞인 기록
- 값이 None이거나 빈 문자열인 경우
- 같은 파일에 계속 누적 기록
- 회차/날짜를 직접 넣는 경우 vs 자동 채움
- 커스텀 헤더로 새 트래커 생성

실행:
    python tests/test_gamba_log_demo.py
"""
import os
import sys

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, HERE)
from gamba_log import log_result

OUT = os.path.join(HERE, "output")


def _clean(path):
    if os.path.exists(path):
        try:
            os.remove(path)
        except PermissionError:
            pass


def _verify_xlsx(path, expected_rows):
    """엑셀 파일을 다시 열어서 실제 데이터 행 수와 내용을 검증한다."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]
    actual_rows = ws.max_row - 1  # 헤더 제외
    ok = actual_rows == expected_rows

    print(f"  [검증] 헤더: {headers}")
    print(f"  [검증] 데이터 행: {actual_rows}행 (기대: {expected_rows}행) {'-> OK' if ok else '-> FAIL'}")
    if not ok:
        return False

    # 마지막 행 내용 출력
    last = {str(headers[i]): ws.cell(ws.max_row, i + 1).value for i in range(len(headers))}
    print(f"  [검증] 마지막 행: {last}")

    # 차트 존재 확인
    has_chart = len(ws._charts) > 0
    print(f"  [검증] 차트: {'있음' if has_chart else '없음'}")
    return ok


def test_1_cold_start():
    """시나리오 1: 파일 없는 상태에서 첫 기록 + 누적 3회"""
    print("\n" + "=" * 60)
    print("[TEST 1] Cold start - 파일 없는 상태에서 첫 기록 + 누적")
    print("=" * 60)

    tracker = os.path.join(OUT, "test1_cold_start.xlsx")
    _clean(tracker)

    log_result({"정확도(%)": 85.0, "SNR(dB)": -10, "오인식(건)": 3, "비고": "baseline"}, tracker=tracker)
    log_result({"정확도(%)": 90.2, "SNR(dB)": -15, "오인식(건)": 1, "비고": "augment"}, tracker=tracker)
    log_result({"정확도(%)": 94.1, "SNR(dB)": -20, "오인식(건)": 0, "비고": "final"}, tracker=tracker)

    return _verify_xlsx(tracker, 3)


def test_2_missing_keys():
    """시나리오 2: 트래커 헤더에는 5개 컬럼이 있지만, 일부 키만 넘기는 경우"""
    print("\n" + "=" * 60)
    print("[TEST 2] Missing keys - 키 일부만 넘기는 불완전한 기록")
    print("=" * 60)

    tracker = os.path.join(OUT, "test2_missing_keys.xlsx")
    _clean(tracker)

    # 1회차: 모든 키 제공 -> 헤더 결정됨
    log_result({"정확도(%)": 85.0, "SNR(dB)": -10, "Loss": 0.42, "오인식(건)": 3, "비고": "full"}, tracker=tracker)
    # 2회차: 정확도만 넘김 (나머지는 빈 칸이어야 함)
    log_result({"정확도(%)": 88.0}, tracker=tracker)
    # 3회차: 비고만 넘김
    log_result({"비고": "memo only"}, tracker=tracker)

    return _verify_xlsx(tracker, 3)


def test_3_extra_keys():
    """시나리오 3: 기존 트래커에 없는 새 키가 포함된 기록 (무시되어야 함)"""
    print("\n" + "=" * 60)
    print("[TEST 3] Extra keys - 트래커에 없는 키 포함")
    print("=" * 60)

    tracker = os.path.join(OUT, "test3_extra_keys.xlsx")
    _clean(tracker)

    # 1회차: 헤더 결정
    log_result({"정확도(%)": 85.0, "SNR(dB)": -10}, tracker=tracker)
    # 2회차: 기존 키 + 트래커에 없는 "새지표" 추가
    log_result({"정확도(%)": 90.0, "SNR(dB)": -15, "새지표": 999, "WER": 0.05}, tracker=tracker)

    return _verify_xlsx(tracker, 2)


def test_4_messy_values():
    """시나리오 4: None, 빈문자열, 문자열이 숫자 위치에, 매우 큰/작은 수"""
    print("\n" + "=" * 60)
    print("[TEST 4] Messy values - 다양한 엣지케이스 값")
    print("=" * 60)

    tracker = os.path.join(OUT, "test4_messy.xlsx")
    _clean(tracker)

    log_result({"정확도(%)": 85.0, "SNR(dB)": -10, "비고": "normal"}, tracker=tracker)
    log_result({"정확도(%)": None, "SNR(dB)": "", "비고": "null test"}, tracker=tracker)
    log_result({"정확도(%)": "N/A", "SNR(dB)": -999.99, "비고": ""}, tracker=tracker)
    log_result({"정확도(%)": 0, "SNR(dB)": 0, "비고": "zero"}, tracker=tracker)
    log_result({"정확도(%)": 100.0, "SNR(dB)": 50, "비고": "perfect"}, tracker=tracker)

    return _verify_xlsx(tracker, 5)


def test_5_manual_round_and_date():
    """시나리오 5: 회차/날짜를 사용자가 직접 지정"""
    print("\n" + "=" * 60)
    print("[TEST 5] Manual round/date - 회차/날짜 직접 지정")
    print("=" * 60)

    tracker = os.path.join(OUT, "test5_manual.xlsx")
    _clean(tracker)

    log_result({"회차": 10, "날짜": "2026-01-15", "정확도(%)": 90.0, "비고": "user set"}, tracker=tracker)
    # 회차/날짜 생략 -> 자동 채움
    log_result({"정확도(%)": 92.0, "비고": "auto fill"}, tracker=tracker)

    ok = _verify_xlsx(tracker, 2)

    # 추가 검증: 1행의 회차가 10인지 확인
    import openpyxl
    wb = openpyxl.load_workbook(tracker)
    ws = wb.active
    r1_round = ws.cell(2, 1).value  # 데이터 1행
    r2_round = ws.cell(3, 1).value  # 데이터 2행
    print(f"  [검증] 1행 회차={r1_round} (기대: 10), 2행 회차={r2_round} (기대: 자동)")
    return ok


def test_6_custom_headers():
    """시나리오 6: 영어 키, 커스텀 헤더, 일반적이지 않은 지표명"""
    print("\n" + "=" * 60)
    print("[TEST 6] Custom headers - 영어/커스텀 헤더")
    print("=" * 60)

    tracker = os.path.join(OUT, "test6_custom.xlsx")
    _clean(tracker)

    log_result({"epoch": 1, "train_loss": 0.52, "val_loss": 0.61, "lr": 0.001, "비고": "start"}, tracker=tracker, sheet="training_log")
    log_result({"epoch": 2, "train_loss": 0.38, "val_loss": 0.45, "lr": 0.0008, "비고": "decay"}, tracker=tracker, sheet="training_log")
    log_result({"epoch": 3, "train_loss": 0.21, "val_loss": 0.30, "lr": 0.0005, "비고": "converging"}, tracker=tracker, sheet="training_log")

    return _verify_xlsx(tracker, 3)


def test_7_rapid_append():
    """시나리오 7: 빠르게 10회 연속 기록 (실제 학습 루프 시뮬레이션)"""
    print("\n" + "=" * 60)
    print("[TEST 7] Rapid append - 10회 연속 빠른 기록")
    print("=" * 60)

    tracker = os.path.join(OUT, "test7_rapid.xlsx")
    _clean(tracker)

    import random
    random.seed(42)
    for i in range(10):
        acc = 80 + i * 1.5 + random.uniform(-0.5, 0.5)
        log_result({
            "정확도(%)": round(acc, 2),
            "SNR(dB)": -10 - i,
            "Loss": round(0.5 - i * 0.04 + random.uniform(-0.01, 0.01), 4),
            "비고": f"epoch_{i+1}"
        }, tracker=tracker)

    return _verify_xlsx(tracker, 10)


def test_8_flexible_inputs():
    """시나리오 8: 딕셔너리 외 다양한 입력 형태 (자연어 문자열, kwargs, PyTorch dict, 리스트)"""
    print("\n" + "=" * 60)
    print("[TEST 8] Flexible inputs - 문자열, kwargs, PyTorch 변수명, 리스트")
    print("=" * 60)

    tracker = os.path.join(OUT, "test8_flexible.xlsx")
    _clean(tracker)

    # 1) 키워드 인자 (kwargs) 형태: log_result(acc=88.5, snr=-10, note="kwargs test")
    log_result(acc=88.5, snr=-10, note="kwargs test", tracker=tracker)

    # 2) PyTorch/WandB 스타일 영문 dict: {"val_acc": 91.2, "train_loss": 0.24, "epoch": 2}
    log_result({"val_acc": 91.2, "val_snr": -15, "loss": 0.24, "epoch": 2, "note": "pytorch dict"}, tracker=tracker)

    # 3) 자연어 / 콘솔 로그 문자열: "3회차 정확도 93.5%, SNR -18dB, 오인식 1건, 자연어 파싱 완료"
    log_result("3회차 정확도 93.5%, SNR -18dB, 오인식 1건, 자연어 파싱 완료", tracker=tracker)

    # 4) 리스트/튜플 형태: [95.0, -22, 0, "list test"]
    log_result([95.0, -22, 0, "list test"], tracker=tracker)

    return _verify_xlsx(tracker, 4)


def test_9_unknown_metrics():
    """시나리오 9: 전혀 생소한 변수명 / 유사한 변수명이 들어왔을 때 (유사도 매핑 & 컬럼 자동 확장)"""
    print("\n" + "=" * 60)
    print("[TEST 9] Unknown metrics - 유사 변수명 매핑 & 새로운 컬럼 자동 확장")
    print("=" * 60)

    tracker = os.path.join(OUT, "test9_unknown.xlsx")
    _clean(tracker)

    # 1) 기본 엑셀 생성: "회차", "날짜", "정확도(%)", "Loss"
    log_result({"정확도(%)": 85.0, "Loss": 0.50, "비고": "baseline"}, tracker=tracker)

    # 2) 유사한 변수명 들어옴: top1_acc (-> 정확도(%)), val_loss (-> Loss)
    log_result({"top1_acc": 88.5, "val_loss": 0.35, "note": "fuzzy match"}, tracker=tracker)

    # 3) 기존 트래커에 전혀 없는 완전히 새로운 지표 들어옴: mAP, BLEU_Score -> 엑셀 컬럼 자동 확장!
    log_result({"top1_acc": 91.0, "val_loss": 0.25, "mAP": 0.78, "BLEU_Score": 32.5, "note": "new metric added"}, tracker=tracker)

    return _verify_xlsx(tracker, 3)


def test_10_zero_arg_caller_search():
    """시나리오 10: log_result()에 인자를 아무것도 안 넘겼을 때 (호출자 스택 프레임의 변수 자동 탐색)"""
    print("\n" + "=" * 60)
    print("[TEST 10] Zero-arg search - 호출자 스코프 local 변수 자동 서치")
    print("=" * 60)

    tracker = os.path.join(OUT, "test10_zero_arg.xlsx")
    _clean(tracker)

    # 호출하는 함수 안의 변수 정의
    epoch = 1
    accuracy = 87.5
    val_loss = 0.35
    note = "stack search test"

    # 인자 없이 1줄만 호출!
    log_result(tracker=tracker)

    return _verify_xlsx(tracker, 1)


def test_11_txt_and_csv_fallback():
    """시나리오 11: .txt 및 .csv 파일 내보내기 (openpyxl 무관한 초경량 텍스트 트래커)"""
    print("\n" + "=" * 60)
    print("[TEST 11] TXT & CSV fallback - 텍스트/CSV 파일 내보내기")
    print("=" * 60)

    txt_tracker = os.path.join(OUT, "test11_py_tracker.txt")
    csv_tracker = os.path.join(OUT, "test11_py_tracker.csv")
    _clean(txt_tracker)
    _clean(csv_tracker)

    # 1) TXT 기록
    log_result({"정확도(%)": 88.5, "SNR(dB)": -10, "비고": "py_txt_1"}, tracker=txt_tracker)
    log_result({"정확도(%)": 92.0, "SNR(dB)": -15, "비고": "py_txt_2"}, tracker=txt_tracker)

    # 2) CSV 기록
    log_result({"정확도(%)": 89.0, "SNR(dB)": -12, "비고": "py_csv_1"}, tracker=csv_tracker)
    log_result({"정확도(%)": 93.5, "SNR(dB)": -18, "비고": "py_csv_2"}, tracker=csv_tracker)

    print(f"  [검증] TXT 파일 존재: {os.path.exists(txt_tracker)}, CSV 파일 존재: {os.path.exists(csv_tracker)}")
    return os.path.exists(txt_tracker) and os.path.exists(csv_tracker)


if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)

    tests = [
        test_1_cold_start,
        test_2_missing_keys,
        test_3_extra_keys,
        test_4_messy_values,
        test_5_manual_round_and_date,
        test_6_custom_headers,
        test_7_rapid_append,
        test_8_flexible_inputs,
        test_9_unknown_metrics,
        test_10_zero_arg_caller_search,
        test_11_txt_and_csv_fallback,
    ]

    results = []
    for t in tests:
        try:
            ok = t()
            results.append((t.__name__, ok))
        except Exception as e:
            print(f"  [ERROR] {e}")
            results.append((t.__name__, False))

    print("\n" + "=" * 60)
    print("[SUMMARY]")
    print("=" * 60)
    for name, ok in results:
        status = "PASS" if ok else "FAIL"
        print(f"  {status}  {name}")
    passed = sum(1 for _, ok in results if ok)
    print(f"\n  {passed}/{len(results)} passed")
    if passed == len(results):
        print("  -> All tests passed!")
    print("=" * 60)
